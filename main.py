from openpyxl.drawing.xdr import XDRPoint2D as Point, XDRPositiveSize2D as Size
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.utils.units import pixels_to_EMU as PixToEMU
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from argparse import ArgumentParser
from time import perf_counter
from base64 import b64decode
from io import BytesIO


# Imagem em convertida para uma string base64.
_img = r'iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAMAAADDpiTIAAAC+lBMVEUAAABkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGRkZGQAjMEAhroAiLwAjsMAjcMAh7wAf7MAdqgAdagAeqwAe64AaZsAaZsAaZsAaZsAeKsAeKsAfbAAgrUAhroAjMAAjcEAjsMAib4AhbkAgLQAdqkAeKsAeawAcqQAd6oAgbQAhLgAjMEAjcIAhroAjMEAiL0AhLgAjsMAi8AAhroAf7MAeasAcqQAap0Ac6UAcqQAaZsAaZsAdqgAfbAAgLQAg7cAjcKCgoKEhISFhYWIiIiKioqMjIyOjo6QkJCTk5OVlZWYmJiampqcnJyenp6hoaGjo6OlpaWoqKiqqqqsrKywsLCysrK2trb/6Gj/6Wj/5V//5l7/5l//30r/3Dr/2Cv/1iD/1yP/1h//2Cn/1R7/2S//2TD/2zf/3D3/3kP/4Er/3kb/4U//5V//5mD/6Wv/6Wv/5V7/5l//6Gb/5mD/2zr/1yj/1R7/1R//1yf/1R7/1h//1yX/1iH/1R7/2Cr/2Cz/1R7/5mD/4lP/3UL/4Ev/5l//5V7/6Wv/52T/6Wv/6Wn/4Er/3Dv/2TNkZGQAaZsAapwAa50AbJ4AbZ8AbaAAbqEAcKIAcaMAcqQAc6UAdKYAdacAdqgAdqkAd6kAd6oAeKoAeKsAeasAeawAeqwAeq0Ae60Ae64AfK8AfbAAfbEAfrEAfrIAf7IAf7MAgLMAgLQAgbQAgbUAgrUAgrYAg7cAhLgAhbkAhroAh7sAiLwAib0Air4Air8Ai8AAjMEAjcL/1R7/1SD/1iL/1iT/1ib/1yf/1yn/1yv/2Cv/2C7/2TD/2TH/2TP/2jT/2jb/2jf/2zn/3Dr/2zz/3D3/3T//3ED/3UH/3kT/3kX/3kb/30f/30n/30r/4Ev/4E3/4E7/4U//4VD/4lL/4lP/4lT/41X/41f/41j/5Fn/5Fr/5Vv/5V3/5l//5mD/52H/5mP/52T/6Gb/6Gf/6Gj/6Wr/6WvxD1Z9AAAAlXRSTlMAECAwQFBgcICPn6+/z9/v7+/f38/Pz8/f7+/v37+vr7+/v7+/r6+vn5+fj4CAcHCAgI+PcGBQUEBAQEBAQFBgcI8wIBAwMAECBAYJCw0PEhQXGRsdICIkJSgqLTE1QDAwIBAwQFBAMCAQYHCAgICAgI+fj4CAcHBgYFBggICPj5+vv8/f3+/v7+/v39/Pz7+vn7+vn+cVdB4AAA/OSURBVHja7MEHAQAhEACg3+82vX0NcsABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAANzPQVhPHisdxPSWudmtByOHwigMw6sC1hqtbSHjqMM0EJvXFtq6dgdnztvC8/1ouOEAiv3///+8Bf2USqU1iG02gnAA6erf71cHmqoqiixJoigIPMcyFEkcnj+9ltdAtZUbAFb5ut42DF0P/OXEnyKI5WI+m50+lMEOAPu7tizT99ey/jRFBv7T6Xh0VgY5AKx6Y7v+7gCy/nzovwz8x8NB/wngALDqft5f8v25yN8bwMTz73XPqgAHgP527O8OIOu/iv2H/X630z6q4gCA9R7561l/NuU/jfxbzQccAKyq23bwAdy+/Txw2DurBddtIAyPMWaXmZmZiTbtY5SZmZnoMKjMzHDZBygzMzMzo6IkthONHCdKHGXr77Jdn/Xm/yTP2JLD8s82AO38725cAGj+N990006VANOKLVv5r1MHgG2X5BvAJP/bWP43LlMJMGJMp4VuOyY0cEZ4mps1818bGPVFsQbwnkz+N9wwaVOAZU6YAA5pYQYkBIpFCIyMNZoFQB2abNLZAFIB2vnfyvK/8YbrJ6oKMLyYOBN4CSABUAJCNADwRioAy391aDGFNQAs/1ta+V83OdcA3Y0IZaIFsAEgHqUAzQYwEQDwBjDN//rrroMJISSUyRYgiH0Ai3ijFIA1gItBi225BpAWgGn+N9D8r4UJgUwDAbwYwA+dEQrQugGwLTRZl28AkgawlX8lQIkCWMTSSG2UArRuAC1WhwZbJg0A1wAm+VcClCgAxK5N9FEK0H4CuOiWddh9HbwBvLWdPxPgGiiPSgAv8kMYqQD4E8CkAUgbwHb+V0N5VALYhNRGKgDyBBBvAJP8KwHKFEAjRB+tAPwTILwBvKGd/1WVAKPHNIBiGM2V67oJIyOzBBBrAG/vzP9amn8lwLQizb9AA3jtNTT/KysBphNpAZjfADIBWP6VANOK1RmrtVm1xSqrLIM0ACz/K/4XAlRsgOTfEOCKSoD/iQCsAUTyvxyGhW7WHDegeI5jar22cikmAD33oIHv2EYfR9mOFzRwHUuHcbENY9uU7TJs32I9pAFk+TcE2DMDDIRR82PSQWhruRv5eq+ggTYOg7QInDa5AuiW4wcMx4IeWJ0nH3tWAd1tL+z6i2tjcGCPzZZIl4BmG4C0AcQbAJb/lTT/ywDg7yyzTxnoQT1PXBOt6WjQewUNtCECxAJYXtSZaF40NnL2kQ15aLWQYHh6yWN/DbYEULgHTPwE6Jokf06Av/6c2dc0YAZERKij8Y5YAN2JCYcjPP1IcO6mWHiPiIjtcrcAPcrvASveALL8uwWg+f/x+2/FJwEjIDnERskCiNMJNMBwiRCnQPw8Xnmz/+Is/4cz+eN7wIQN4BU0/04BaPyN/H/79WQohOaSfGKjZAGcWDgdIQZoufr62BHs9FQwYI9FHi26B0zYAFzWLQAb/jT/ny/eGwpghB2fsGubDWwv5gwoSQA/JGICPs2w4/Rrpul4EeIMfsGIPNvUADSz5nMzx+ip0/FffA+YOH8mQGf+v/z8888/zSmSfywo+rXMOAw1aQGKt4Epse9QFy03zksmSH/c0bGaJuwSxssMdEM4E+pQAms9WnwPGP8EiOXfJQCLvzH90/h//PGHk/vJ39PFc4M7BgF8K3VRnIwnmOzNGJ/P9SjnD47KvQhsTfPvaADEe8DEDWAiQHb4N/P//vuZgICPnwBRXks/RKNMAfh0bFEyFmlTE18abOzXhAZwaFGpU8Dij8o3gFkBkuqPxU/z/+7bPQsLkP/eHhKUKECiI26AhipayykOYh35NWaPCdGGUbOlsAEU7wHjG4BEgKT6aw9/mv83x0oJkG2wzBIFiCzg8NGhnvxXPzdNv6gA4HBXvdEhfgmIeA8Ynj8TIFv9NYf/N19/dYmkAOkIC8oTwNWARyd5acZafpqmWADxHzxi6o8+Kt0AZgXoqP5a+X/5xf5SAmSnAL0sAVxA8ZKwkbN3ACVK0iwqAHilCbC1bAPYJUC2+qPxf/vNVzT/z4+WFMBIZ95SBBCHafG3DnRsAsArB6OoAFZpAmyGN4DiPWB4A8BYHmCvjuqPDX+a/2fzJQVIB1E4ZgGAD9Pt2bPFybRSVACjNAHWyGkAHyzeADJWAjils/pjw/+zzz79RFaAtM/WxyxAkETHyWn1PPu4qABQmgBJ/lgDULwBZOwMMK+j+mP50/g//ng/SQHstC8aswBed3QGf1EQXzcM1QTYo2ADeKu4AUxYvg5wcVf19znL/6MPD5MUIP2UvTEL4HRHVyuQFWnjqCbAllSAAfaA8Q0AZUOAk7urv88++6SR/wfHSAoApE2gmgA+0jeI73SqJsCmNP++94Dh+a9QB5jZXf19yuL/4P35sgKEpI1qAkT4Efh1I1JNgDW4BvChQRpAyvK7ARzPVX8s//ffe/dSWQGCtApUTACCZ4kfRJQTgGsA0D1geAPYXQHuzVd/rfzfeUdWAD/9nNUSQEcFEFeBiglQtAHkGwB+/O87E6n+aPyN/N/eX1IAR1UBzHwBuB9STACa/+BLAFN2rdP8ZyHVHxv+NP83hyeAo6oA2iQKgDYADxZaApiw4i5TALD3LKz6Y/m/9eabr582/QWAiRQg2wDgDeAy6++8Q8KOHHVgnHQxXv2x4f/mG68doaIAlQBT9AKQ3wAW+z6Yk2cJqz86/Gn+igpQXQJ6NICrTEFP9j31uLzqj07/r7322qtqClAVgfl7wDYCyj4nzp3zdxfcuj9h9ceG/6uvvFwJoKYAuUsAdwaAveYi4bN1f+nCj9zqr5n/S4dPVwGgv/sAmmIC4A1gmv8JaPz8uj9x9Ufjp/m/KNsFeKSNpaoAtUm8E7gk3gC25/95eP5/cuv+xNXfq434X3xB9j5AoOSNIH5BGI6j7MOg1YVLAFfB8ud3feRXf638X3j+uQMlBQiVfRbg4Vnhf5+nmgCbC5cATrH5X776e4nl/9zQHgcTZdcDxCAkSi4TqgmwsagBXA9gn6FUfzT+55979gJJAQx11wOkp2aAAE3dFUHbivaA7QYwbxjVHxv+zzx9jqQANdLG5aMwxioAP7zFTUCk3ppAwR6wZQD2GU71x/J/SnZJmJ82AUgU4xXAxTYA49saXPUEWArfA7Y+wInDqf6ee5bm/+RBUgJk94ciUdjjFcDg6lNhCWCoJ8C6+B6wnQHmSld/reFP8z8f5ASwSRsPicIbrwAQ9jjGTNNUT4At8T1g2wPMHkb1x4b/E4+fLSlA2gSaSBSxNl4B7B4n4ieHKCjAFL4EcAeA4VR/LP/HDpITwCZtIvQGuzOgAIGsAPi6UPEuHwUFgKWyT4Da+TMBJKu/9vCn+Z8PMgJkKwAbba5iHVDcHgKEkgLwKhp505eppACbIEsAEwHkqz8a/2OPnVlYgNBE/y+emI+8PQh/DY/oH9SGIwD4/InwB7igpABT6BLARADJ6o/mTzm4mAD4K3M0nyQYWFj4m1a07FveRDODIykAP0f5wstXpKkpAKyL7QGjAqDV3w99Vn+PNTgHigtA8cyOREKS4IgPix0NErhXfIquyrExHAHAFE1GLverlBNgCtsDRgXghv8g1R/jwL4EoESuCQzNDvLezqdnU45dM33dNOlE+PsiQ1IAvk6NbfTdwTaoKgCsj+wBowJ0V38/DlD9Mc6F4gKkhEEQRL3ez2kiBwWER1SYUzxrKAKAnVHRs5tvucycvw3qClBfit8DRgVIhr9E9dfgooMHEIAHL/TsmBSBP65TG1deALCJmNgChQWA3Zbj9oAxAaSrP8ah0IcAyZxZ9A3NRp4BHiIAHtcQBAAzFrprgNICwG7Ldu0BYwLQ+OWqP8bZ0I8AncU7hc+DbxJwopoOYgHAjOQF4PtOjNgBUFwAqK/ctQeMCiBd/TEOg/4EANDssL9vTzCxCsIxshs3AYGqJiuA+P3vXHuitgCU7VfueAkIFUC6+qOcfTD0LQDFcMK+vj/FcMPMj/uO2b32FnAsN+j+Ad1pYfb6Ghr8pDTbz3gVdZWY/K/RRZ61sKE06ttvmDIFIKz+ZhxNOarBkQ2OaHB4g9MZZyQceAjAIAIwTNtpYJtQDLMJPs56HajDMNHMhiU104SJRlj9zQdpxAJIgwtQ0T/C6m/yBAigon+E1d8M4ew/nQSoGKD6U00Ai3tCU1GcAe79qSaAI/HVOxWt6u/rPu79qSaAy63XrihOWv19XvDen2oCJP+sCRX98/333/Z1709BAWJu4U9FcVrDv+i9PwUFMPiVpBXFKVT9XXj6gQeedRGNX0UBajLfvVZRpPo77xCgHHSRmgJEMt+8VVGg+rvwEGAcpKQADrJdo6I4Baq/s6DFOQoKYMp9+2bFjN7V34HQ4gz1BDBjuS/frJjf+95flwAXKSRATfYbuCuO7n3vL9nxc16y/l8NAcyAyC6pqdin572/ZMHvgcnyTyUEsAPxWuKK4hzb+97fecyAA5tt4PmggABmzY/JUPKv2OfS3vf+Ljr70EPPfqzJAeMVwLAcLySduFAhwQHvvSPe9cFxNoxTAJ3wRCZUSHEaVv0J8x+rABAhy/ErZNn/Urz64zkDZPDlBfC6Rr+tQYU8hyzAqj+O8w8EOSwvlhTAyqbvGlAxJPY/l6v+uPgPhSFg+1ICaO3w/VqV/nA5aOGFedXfOQfAkNDsAAbHdRxnRHsxKg446wJ0+j//7EMPqT6d/9izawMJghgIgJvdMuOQ+RjDEwX+0o2OM9B2WQdm9/A+PD+8/fxdZv/z/vqcwM48P0aIHgAAAAAAAECxNE0zkeeZoB8T0C6lxIuiKFl1oWT0B/UBNVCdPede103Tku6kJU1T19wEtEClNIZP2XPwfd8PwziO0wF9GIah77kI1IJYgiwBffFz+pQ9BT/Py7KuW7QuyzxTEagE3AGpQJqAEll+Hf88U/TGWiesNWZb5/mqApgEdI3/igsg+S9rjN/7QLw/VGBbl2MDmqaqMAconQG4ABcNYKf8TwWIM0CegA5pfrcGLAuvAkJ2ATcrACYAjbtA6cDAJ4D5hE8Cg6Svcw8IacYdkJPg4Q6gvxDvAk43AbgI0H4PWFU1aURNqmond4HADwHchLw4oS85/fjfnh0LAAAAAAzyt57FrtLIBgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAALALx5Epwd57JLAAAAABJRU5ErkJggg=='


if __name__ == '__main__':
    runtime = perf_counter()
    parser = ArgumentParser()
    parser.add_argument('dir')
    args = parser.parse_args()

    if args.dir[args.dir.rfind('.'):] != '.xlsx':
        raise TypeError(f'Apenas arquivos Excel são aceitos, mas foi passado um arquivo com extensão {args.dir[args.dir.rfind("."):]}')
    
    else:
        output = args.dir.replace('.xlsx', ' - Formatado.xlsx')

    wb = load_workbook(args.dir)  # Carrega a planilha na memória.
    sheet = wb.worksheets[0]  # Carrega a primeira parte da planilha na memória.
    row_count = sheet.max_row + 1  # Da linha que será adicionada.
    column_count = sheet.max_column - 2  # Das duas colunas que serão excluídas.
    thin = Side(border_style='thin', color='000000')  # Cria o objeto dos lados da borda.

    sheet.sheet_view.showGridLines = False  # Retira as linhas de grade
    sheet.delete_cols(5)  # Remove a coluna E
    sheet.delete_cols(2)  # Remove a coluna B

    sheet[f'A{row_count}'] = 'TOTAL:'  # Isso tem que vir primeiro por essa célula ser referenciada posteriormente.
    sheet.merge_cells(f'A{row_count}:C{row_count}')  # Merge as células da última linha das colunas A a C.

    sheet.row_dimensions[1].height = 53  # Ajusta a altura da primeira linha.
    sheet.column_dimensions['A'].width = 56  # Ajusta a largura da primeira linha.

    for col in sheet.iter_cols():  # retorna todas as colunas entre A e a última.
        start_time = perf_counter()
        print(f'Formatando a coluna {get_column_letter(col[0].column)}... ', end = '')

        for cell in col:  # retorna todas as células da coluna.
            cell.border = Border(thin, thin, thin, thin)  # Coloca todas as bordas na célula.
            cell.alignment = Alignment('center', 'center')  # Alinha o conteúdo da célula nos centros.

        col[0].fill = PatternFill('solid', start_color='002060')  # Preenche de azul o fundo da primeira célula (col[0])
        col[row_count - 1].fill = PatternFill('solid', start_color='002060')  # Preenche de azul o fundo da última célula (col[row_count - 1])
        col[row_count - 1].font = Font(color='FFFFFF', bold=True)  # Põe o texto da última célula em negrito e branco.

        if col[0].column == 2:  # Ajustando a coluna B
            max_width = max(len(str(cell.value)) * 0.89 for cell in col)  # Encontra a maior célula qual vai definir o tamanho da coluna.
            sheet.column_dimensions['B'].width = max_width  # Ajusta o tamanho da coluna.

        elif col[0].column == 3:  # Ajustando a coluna C
            max_width = max(len(str(cell.value)) * 0.89 for cell in col)  # Idem supra.
            sheet.column_dimensions['C'].width = max_width  # Idem supra.

        elif col[0].column > 3:  # Ajusta o tamanho e conteúdo das colunas D em diante.
            col[row_count - 1].value = f'=COUNTA({get_column_letter(col[0].column)}2:' \
                                       f'{get_column_letter(col[0].column)}{row_count - 1})'  # Coloca a fórmula
            sheet.column_dimensions[get_column_letter(col[0].column)].width = 20  # Ajusta o tamanho
        
        print(f'coluna formatada em {perf_counter() - start_time:.2f} segundos.')

    byte_image = b64decode(_img)  # Decodifica o texto base64 para bytes.
    image = Image(BytesIO(byte_image))  # Interpreta os bytes como imagem e carrega-a na memória.
    image.anchor = AbsoluteAnchor(Point(PixToEMU(0), PixToEMU(8)),  # Ajusta a posição da imagem.
                                  Size(PixToEMU(204), PixToEMU(55)))  # Ajusta o tamanho da imagem.
    sheet.add_image(image)  # Carrega a imagem na folha da planilha.

    wb.save(output)  # Salva a planilha no disco e a retira da memória.
    print(f'Programa executado com êxito em {perf_counter() - runtime:.2f} segundos.')
